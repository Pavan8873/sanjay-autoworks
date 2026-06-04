from datetime import date, timedelta
from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Q
from django.shortcuts import render
from django.http import HttpResponse
from django.conf import settings

from customers.models import Customer, Vehicle
from jobcards.models import JobCard
from billing.models import Invoice
from inventory.models import Part
from reminders.models import Reminder


@login_required
def dashboard(request):
    today = date.today()
    upcoming = today + timedelta(days=30)

    open_jobs = JobCard.objects.exclude(status__in=["billed", "cancelled"]).select_related("customer", "vehicle").order_by("-created_at")[:10]
    pending_payments = Invoice.objects.filter(paid=False).select_related("jobcard__customer")[:10]
    upcoming_reminders = Reminder.objects.filter(status="pending", due_date__lte=upcoming).select_related("vehicle__customer")[:10]

    stats = {
        "customers": Customer.objects.count(),
        "vehicles": Vehicle.objects.count(),
        "open_jobs": JobCard.objects.exclude(status__in=["billed", "cancelled"]).count(),
        "completed_today": JobCard.objects.filter(completed_at__date=today).count(),
        "revenue_month": Invoice.objects.filter(paid=True, created_at__year=today.year, created_at__month=today.month).aggregate(s=Sum("total"))["s"] or 0,
        "pending_amount": Invoice.objects.filter(paid=False).aggregate(s=Sum("total"))["s"] or 0,
        "parts_count": Part.objects.count(),
        "low_stock_count": Part.objects.filter(stock__lte=5).count(),
    }

    return render(request, "core/dashboard.html", {
        "stats": stats,
        "open_jobs": open_jobs,
        "pending_payments": pending_payments,
        "low_stock": [p for p in Part.objects.all() if p.low_stock][:10],
        "upcoming_reminders": upcoming_reminders,
    })


def _parse_report_period(request):
    """Return (start_date, end_date, period_key, label) from request."""
    today = date.today()
    period = request.GET.get("period", "monthly")

    # Custom date range overrides period tabs
    from_str = request.GET.get("from_date", "")
    to_str = request.GET.get("to_date", "")
    if from_str and to_str:
        try:
            start_date = date.fromisoformat(from_str)
            end_date = date.fromisoformat(to_str)
            if start_date > end_date:
                start_date, end_date = end_date, start_date
            return start_date, end_date, "custom", f"Custom: {start_date.strftime('%d %b')} – {end_date.strftime('%d %b %Y')}"
        except ValueError:
            pass

    if period == "today":
        return today, today, "today", f"Today — {today.strftime('%d %b %Y')}"
    if period == "weekly":
        return today - timedelta(days=6), today, "weekly", "Last 7 Days"
    if period == "yearly":
        return date(today.year, 1, 1), today, "yearly", f"Year {today.year}"
    # default: monthly
    return date(today.year, today.month, 1), today, "monthly", today.strftime("%B %Y")


@login_required
def reports(request):
    start_date, end_date, period, label = _parse_report_period(request)

    invoices = Invoice.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
    )
    jobs = JobCard.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
    )

    stats = {
        "revenue_paid": invoices.filter(paid=True).aggregate(s=Sum("total"))["s"] or 0,
        "revenue_all": invoices.aggregate(s=Sum("total"))["s"] or 0,
        "pending_amount": invoices.filter(paid=False).aggregate(s=Sum("total"))["s"] or 0,
        "invoice_count": invoices.count(),
        "paid_count": invoices.filter(paid=True).count(),
        "pending_count": invoices.filter(paid=False).count(),
        "jobs_created": jobs.count(),
        "jobs_completed": jobs.filter(status__in=["completed", "billed"]).count(),
        "new_customers": Customer.objects.filter(
            created_at__date__gte=start_date, created_at__date__lte=end_date
        ).count(),
        "new_vehicles": Vehicle.objects.filter(
            created_at__date__gte=start_date, created_at__date__lte=end_date
        ).count(),
    }

    # Payment method breakdown
    payment_breakdown = (
        invoices.filter(paid=True)
        .values("payment_method")
        .annotate(count=Count("id"), total=Sum("total"))
        .order_by("-total")
    )

    # Recent invoices
    recent_invoices = (
        invoices.select_related("jobcard__customer", "jobcard__vehicle")
        .order_by("-created_at")
    )

    # All jobs in period
    recent_jobs = (
        jobs.select_related("customer", "vehicle")
        .order_by("-created_at")
    )

    return render(request, "core/reports.html", {
        "period": period,
        "label": label,
        "start_date": start_date,
        "end_date": end_date,
        "stats": stats,
        "recent_invoices": recent_invoices,
        "recent_jobs": recent_jobs,
        "payment_breakdown": payment_breakdown,
        "from_date": request.GET.get("from_date", ""),
        "to_date": request.GET.get("to_date", ""),
    })


@login_required
def reports_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("openpyxl not installed. Run: pip install openpyxl", status=500)

    start_date, end_date, period, label = _parse_report_period(request)
    today = date.today()

    invoices = Invoice.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
    ).select_related("jobcard__customer", "jobcard__vehicle")

    jobs = JobCard.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
    ).select_related("customer", "vehicle")

    shop_name = getattr(settings, "SHOP_NAME", "AutoCare Service Center")

    wb = openpyxl.Workbook()

    # ── Styles ───────────────────────────────────────────────
    hdr_fill   = PatternFill("solid", fgColor="0F766E")
    hdr2_fill  = PatternFill("solid", fgColor="1D4ED8")
    sum_fill   = PatternFill("solid", fgColor="F0FDF4")
    alt_fill   = PatternFill("solid", fgColor="F8FAFC")
    hdr_font   = Font(bold=True, color="FFFFFF", size=10)
    title_font = Font(bold=True, size=13, color="0F766E")
    sub_font   = Font(size=9, color="64748B")
    bold10     = Font(bold=True, size=10)
    thin       = Side(style="thin", color="E2E8F0")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    center     = Alignment(horizontal="center", vertical="center")
    right      = Alignment(horizontal="right", vertical="center")

    def set_hdr(ws, row, cols, fill=None):
        fill = fill or hdr_fill
        for c in cols:
            cell = ws.cell(row=row, column=c)
            cell.fill = fill
            cell.font = hdr_font
            cell.alignment = center
            cell.border = border

    def style_data_row(ws, row, ncols, alt=False):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            if alt:
                cell.fill = alt_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    # ── Sheet 1: Summary ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 22

    # Title block
    ws1["A1"] = shop_name
    ws1["A1"].font = title_font
    ws1["A2"] = f"Report: {label}"
    ws1["A2"].font = Font(bold=True, size=11)
    ws1["A3"] = f"Period: {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}    |    Generated: {today.strftime('%d %b %Y')}"
    ws1["A3"].font = sub_font
    ws1.merge_cells("A1:B1")
    ws1.merge_cells("A2:B2")
    ws1.merge_cells("A3:B3")

    ws1.append([])
    hdr_row = 5
    ws1.append(["Metric", "Value"])
    set_hdr(ws1, hdr_row, [1, 2])

    summary_data = [
        ("Revenue Collected (Paid)", f"Rs {(Invoice.objects.filter(created_at__date__gte=start_date, created_at__date__lte=end_date, paid=True).aggregate(s=Sum('total'))['s'] or 0):,.2f}"),
        ("Revenue (All Invoices)",   f"Rs {(Invoice.objects.filter(created_at__date__gte=start_date, created_at__date__lte=end_date).aggregate(s=Sum('total'))['s'] or 0):,.2f}"),
        ("Pending Amount",           f"Rs {(Invoice.objects.filter(created_at__date__gte=start_date, created_at__date__lte=end_date, paid=False).aggregate(s=Sum('total'))['s'] or 0):,.2f}"),
        ("Total Invoices",           invoices.count()),
        ("Paid Invoices",            invoices.filter(paid=True).count()),
        ("Pending Invoices",         invoices.filter(paid=False).count()),
        ("Jobs Created",             jobs.count()),
        ("Jobs Completed / Billed",  jobs.filter(status__in=["completed", "billed"]).count()),
        ("New Customers",            Customer.objects.filter(created_at__date__gte=start_date, created_at__date__lte=end_date).count()),
        ("New Vehicles",             Vehicle.objects.filter(created_at__date__gte=start_date, created_at__date__lte=end_date).count()),
    ]
    for i, (metric, value) in enumerate(summary_data):
        r = hdr_row + 1 + i
        ws1.cell(row=r, column=1, value=metric)
        ws1.cell(row=r, column=2, value=value).alignment = right
        style_data_row(ws1, r, 2, alt=(i % 2 == 1))

    # ── Sheet 2: Invoices ─────────────────────────────────────
    ws2 = wb.create_sheet("Invoices")
    for col, width in zip("ABCDEFGH", [16, 16, 28, 18, 14, 16, 18, 12]):
        ws2.column_dimensions[col].width = width

    ws2["A1"] = shop_name
    ws2["A1"].font = title_font
    ws2["A2"] = f"Invoices — {label}  |  Generated: {today.strftime('%d %b %Y')}"
    ws2["A2"].font = sub_font
    ws2.merge_cells("A1:H1")
    ws2.merge_cells("A2:H2")
    ws2.append([])

    inv_headers = ["Invoice #", "Job Card #", "Customer", "Vehicle", "Date", "Amount (Rs)", "Payment Method", "Status"]
    hdr_row2 = 4
    ws2.append(inv_headers)
    set_hdr(ws2, hdr_row2, range(1, 9), fill=hdr2_fill)

    for i, inv in enumerate(invoices.order_by("-created_at")):
        r = hdr_row2 + 1 + i
        ws2.cell(row=r, column=1, value=inv.invoice_number)
        ws2.cell(row=r, column=2, value=inv.jobcard.job_number)
        ws2.cell(row=r, column=3, value=inv.jobcard.customer.name)
        ws2.cell(row=r, column=4, value=inv.jobcard.vehicle.registration_number)
        ws2.cell(row=r, column=5, value=inv.created_at.strftime("%d %b %Y"))
        amt = ws2.cell(row=r, column=6, value=float(inv.total))
        amt.number_format = '#,##0.00'
        amt.alignment = right
        ws2.cell(row=r, column=7, value=inv.payment_method or "—")
        status_cell = ws2.cell(row=r, column=8, value="Paid" if inv.paid else "Pending")
        if inv.paid:
            status_cell.font = Font(color="15803D", bold=True, size=9)
        else:
            status_cell.font = Font(color="B45309", bold=True, size=9)
        style_data_row(ws2, r, 8, alt=(i % 2 == 1))

    if not invoices.exists():
        ws2.append(["—", "—", "No invoices in this period", "—", "—", "—", "—", "—"])

    # ── Sheet 3: Job Cards ────────────────────────────────────
    ws3 = wb.create_sheet("Job Cards")
    for col, width in zip("ABCDEF", [16, 28, 18, 14, 16, 18]):
        ws3.column_dimensions[col].width = width

    ws3["A1"] = shop_name
    ws3["A1"].font = title_font
    ws3["A2"] = f"Job Cards — {label}  |  Generated: {today.strftime('%d %b %Y')}"
    ws3["A2"].font = sub_font
    ws3.merge_cells("A1:F1")
    ws3.merge_cells("A2:F2")
    ws3.append([])

    job_headers = ["Job Card #", "Customer", "Vehicle", "Date", "Advisor", "Status"]
    hdr_row3 = 4
    ws3.append(job_headers)
    set_hdr(ws3, hdr_row3, range(1, 7))

    for i, job in enumerate(jobs.order_by("-created_at")):
        r = hdr_row3 + 1 + i
        ws3.cell(row=r, column=1, value=job.job_number)
        ws3.cell(row=r, column=2, value=job.customer.name)
        ws3.cell(row=r, column=3, value=job.vehicle.registration_number)
        ws3.cell(row=r, column=4, value=job.created_at.strftime("%d %b %Y"))
        ws3.cell(row=r, column=5, value=job.service_advisor or "—")
        ws3.cell(row=r, column=6, value=job.get_status_display())
        style_data_row(ws3, r, 6, alt=(i % 2 == 1))

    if not jobs.exists():
        ws3.append(["—", "No job cards in this period", "—", "—", "—", "—"])

    # ── Serve ─────────────────────────────────────────────────
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    fname = f"report_{period}_{today.isoformat()}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    wb.save(response)
    return response


@login_required
def reports_pdf(request):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    except ImportError:
        return HttpResponse("PDF library not installed. Run: pip install reportlab", status=500)

    start_date, end_date, period, label = _parse_report_period(request)
    today = date.today()

    invoices = Invoice.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
    )
    jobs = JobCard.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
    )

    stats = {
        "Revenue (Paid)": invoices.filter(paid=True).aggregate(s=Sum("total"))["s"] or 0,
        "Revenue (All Invoices)": invoices.aggregate(s=Sum("total"))["s"] or 0,
        "Pending Amount": invoices.filter(paid=False).aggregate(s=Sum("total"))["s"] or 0,
        "Total Invoices": invoices.count(),
        "Paid Invoices": invoices.filter(paid=True).count(),
        "Pending Invoices": invoices.filter(paid=False).count(),
        "Jobs Created": jobs.count(),
        "Jobs Completed / Billed": jobs.filter(status__in=["completed", "billed"]).count(),
        "New Customers": Customer.objects.filter(
            created_at__date__gte=start_date, created_at__date__lte=end_date
        ).count(),
        "New Vehicles": Vehicle.objects.filter(
            created_at__date__gte=start_date, created_at__date__lte=end_date
        ).count(),
    }

    recent_invoices = (
        invoices.select_related("jobcard__customer", "jobcard__vehicle")
        .order_by("-created_at")
    )

    response = HttpResponse(content_type="application/pdf")
    fname = f"report_{period}_{today.isoformat()}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{fname}"'

    doc = SimpleDocTemplate(
        response, pagesize=A4,
        rightMargin=14 * mm, leftMargin=14 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    PRIMARY = colors.HexColor("#0f766e")
    BLUE = colors.HexColor("#1d4ed8")
    MUTED = colors.HexColor("#475569")
    LIGHT = colors.HexColor("#f8fafc")
    LINE = colors.HexColor("#e2e8f0")

    title_style = ParagraphStyle("ts", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=17, textColor=PRIMARY, spaceAfter=3)
    sub_style = ParagraphStyle("ss", parent=styles["Normal"], fontSize=8.5, textColor=MUTED)
    section_style = ParagraphStyle("sec", parent=styles["Heading3"], fontSize=11, textColor=colors.HexColor("#0f172a"), spaceBefore=10, spaceAfter=5)

    shop_name = getattr(settings, "SHOP_NAME", "AutoCare Service Center")
    shop_addr = getattr(settings, "SHOP_ADDRESS", "")
    shop_phone = getattr(settings, "SHOP_PHONE", "")
    shop_gstin = getattr(settings, "SHOP_GSTIN", "")

    story = []
    story.append(Paragraph(shop_name, ParagraphStyle("sn", parent=styles["Normal"], fontSize=11, fontName="Helvetica-Bold", textColor=PRIMARY)))
    if shop_addr:
        story.append(Paragraph(shop_addr, sub_style))
    if shop_phone or shop_gstin:
        story.append(Paragraph(f"Ph: {shop_phone}  |  GSTIN: {shop_gstin}", sub_style))
    story.append(Spacer(1, 6))
    story.append(HRFlowable(width="100%", thickness=1.5, color=PRIMARY))
    story.append(Spacer(1, 4))
    story.append(Paragraph(label, title_style))
    story.append(Paragraph(f"Period: {start_date.strftime('%d %b %Y')}  to  {end_date.strftime('%d %b %Y')}    |    Generated: {today.strftime('%d %b %Y')}", sub_style))
    story.append(Spacer(1, 10))

    # Summary table
    story.append(Paragraph("Summary", section_style))
    summary_rows = [["Metric", "Value"]]
    for k, v in stats.items():
        if isinstance(v, (Decimal, float, int)) and ("Revenue" in k or "Amount" in k):
            value = f"Rs {Decimal(str(v)):,.2f}"
        else:
            value = str(v)
        summary_rows.append([k, value])

    summary_tbl = Table(summary_rows, colWidths=[110 * mm, 65 * mm])
    summary_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), PRIMARY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.4, LINE),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [LIGHT, colors.white]),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(summary_tbl)

    # Invoice list
    story.append(Paragraph("Invoice Details", section_style))
    inv_rows = [["Invoice #", "Job Card", "Customer", "Vehicle", "Date", "Amount", "Status"]]
    for inv in recent_invoices:
        inv_rows.append([
            inv.invoice_number,
            inv.jobcard.job_number,
            inv.jobcard.customer.name,
            inv.jobcard.vehicle.registration_number,
            inv.created_at.strftime("%d %b %Y"),
            f"Rs {inv.total:,.2f}",
            "Paid" if inv.paid else "Pending",
        ])
    if len(inv_rows) == 1:
        inv_rows.append(["—", "—", "No invoices in this period", "—", "—", "—", "—"])

    inv_tbl = Table(inv_rows, colWidths=[24*mm, 24*mm, 40*mm, 22*mm, 22*mm, 24*mm, 19*mm])
    inv_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BLUE),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.35, LINE),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [LIGHT, colors.white]),
        ("ALIGN", (5, 0), (5, -1), "RIGHT"),
        ("ALIGN", (6, 0), (6, -1), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(inv_tbl)

    story.append(Spacer(1, 10))
    story.append(HRFlowable(width="100%", thickness=0.5, color=LINE))
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"Generated by AutoCare Service Manager  ·  {today.strftime('%d %b %Y')}", sub_style))

    doc.build(story)
    return response

